"""
이터널리턴 패치노트 → TC(테스트 케이스) 변환기
- 입력: patchnote_changes.csv  (스크립트와 같은 폴더)
- 출력: patchnote_TC.xlsx      (스크립트와 같은 폴더)
- 컬럼: TC ID | 테스트 항목 | 사전조건 | 테스트 절차 | 기댓값 | 결과
- 대상: 실험체 스킬 / 실험체 기본 스탯 유형만 포함
"""

import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

_DIR       = Path(__file__).parent
INPUT_CSV  = _DIR / "patchnote_changes.csv"
OUTPUT_XLS = _DIR / "patchnote_TC.xlsx"

ARMOR_SLOTS = frozenset({"옷", "모자", "머리", "팔", "발", "목", "장갑", "허리", "반지", "귀걸이", "보조"})

# 실험체(캐릭터) 화이트리스트 — 이 목록에 없는 캐릭터 칸은 TC 생성에서 제외
CHARACTER_WHITELIST = frozenset({
    "가넷", "나딘", "나타폰", "다니엘", "다르코", "데비&마를렌", "띠아",
    "라우라", "레녹스", "레니", "레온", "로지", "르노어", "리 다이린",
    "마르티나", "마이", "마커스", "매그너스", "미르카",
    "바냐", "바바라", "버니스", "비앙카", "블레어",
    "샬럿", "셀린", "쇼우", "쇼이치", "수아", "슈린", "시셀라", "실비아",
    "아델라", "아디나", "아드리아나", "아르다", "아비게일", "아야", "아이솔", "아이작",
    "알렉스", "알론소", "얀", "에스텔", "에이든", "에키온", "엘레나", "엠마",
    "유민", "유스티나", "유키", "이렘", "이바", "이슈트반", "이안", "일레븐",
    "자히르", "재키", "제니",
    "카밀로", "카티야", "캐시", "칼라", "케네스", "코렐라인", "클로에", "키아라",
    "타지아", "테오도르",
    "펜리르", "펠릭스", "프리야", "피오라", "피올로",
    "하트", "헤이즈", "헨리", "현우", "혜진", "히스이",
    "니아", "루크", "윌리엄",
    "츠바메",
})

# ── 유형 분류 ──────────────────────────────────────────────────────────────────

def classify(row: dict) -> str:
    char  = row["캐릭터"]
    skill = row["스킬"]
    if "인퓨전" in char:
        return "cobalt_infusion"
    if "사용 시" in skill:
        return "cobalt_mode"
    if skill in ARMOR_SLOTS:
        return "armor"
    if re.search(r"\((?:[QWERPD]|과전하|패시브)", skill):
        tc_type = "char_skill"
    else:
        tc_type = "char_stat"
    # 캐릭터 화이트리스트에 없는 항목은 제외
    if char not in CHARACTER_WHITELIST:
        return "non_char"
    return tc_type


# ── TC 필드 생성 ───────────────────────────────────────────────────────────────

def _expected(row: dict, label: str = "") -> str:
    prefix = f"{label} " if label else ""
    return (
        f"{prefix}{row['항목']}: {row['변경후']}\n"
        f"(변경 전: {row['변경전']})"
    )


def build_tc_char_skill(row: dict, version: str) -> dict:
    char, skill, item = row["캐릭터"], row["스킬"], row["항목"]
    return {
        "테스트 항목": f"[실험체] {char} - {skill} {item} 수치 확인",
        "사전조건": (
            f"· {version} 패치 적용 완료\n"
            f"· {char} 실험체 선택 가능 상태"
        ),
        "테스트 절차": (
            "1. 게임 클라이언트 실행 및 패치 버전 확인\n"
            f"2. 훈련 모드 진입 후 {char} 선택\n"
            f"3. {skill} 스킬 레벨 최대화 후 스킬 정보창 확인\n"
            f"4. {item} 수치가 패치 내용과 일치하는지 검증"
        ),
        "기댓값": _expected(row, skill),
    }


def build_tc_char_stat(row: dict, version: str) -> dict:
    char, item = row["캐릭터"], row["항목"]
    return {
        "테스트 항목": f"[실험체] {char} - {item} 수치 확인",
        "사전조건": (
            f"· {version} 패치 적용 완료\n"
            f"· {char} 실험체 선택 가능 상태"
        ),
        "테스트 절차": (
            "1. 게임 클라이언트 실행 및 패치 버전 확인\n"
            f"2. 훈련 모드 진입 후 {char} 선택\n"
            f"3. 캐릭터 정보창(스탯창)에서 {item} 확인\n"
            "4. 수치가 패치 내용과 일치하는지 검증"
        ),
        "기댓값": _expected(row),
    }


def build_tc_armor(row: dict, version: str) -> dict:
    item_name, slot, stat = row["캐릭터"], row["스킬"], row["항목"]
    slot_label = f" ({slot})" if slot else ""
    return {
        "테스트 항목": f"[방어구] {item_name}{slot_label} - {stat} 수치 확인",
        "사전조건": (
            f"· {version} 패치 적용 완료\n"
            "· 게임 내 아이템 정보 조회 가능 상태"
        ),
        "테스트 절차": (
            "1. 게임 클라이언트 실행 및 패치 버전 확인\n"
            "2. 훈련 모드 진입 후 제작 화면 접근\n"
            f"3. {slot_label.strip(' ()')} 슬롯에서 {item_name} 아이템 검색\n"
            f"4. 아이템 정보창에서 {stat} 수치 확인\n"
            "5. 수치가 패치 내용과 일치하는지 검증"
        ),
        "기댓값": _expected(row, item_name),
    }


def build_tc_cobalt_mode(row: dict, version: str) -> dict:
    char, weapon_cond, item = row["캐릭터"], row["스킬"], row["항목"]
    # "돌격 소총 사용 시" → 무기 이름만 추출
    weapon = re.sub(r"\s*사용\s*시$", "", weapon_cond).strip()
    return {
        "테스트 항목": f"[코발트 프로토콜] {char} - {weapon_cond} {item} 확인",
        "사전조건": (
            f"· {version} 패치 적용 완료\n"
            "· 코발트 프로토콜 모드 진행 가능 상태"
        ),
        "테스트 절차": (
            "1. 게임 클라이언트 실행 및 패치 버전 확인\n"
            "2. 코발트 프로토콜 모드 게임 시작\n"
            f"3. {char} 실험체 선택\n"
            f"4. 인게임에서 {weapon} 장착 후 전투 진행\n"
            f"5. {item} 수치 확인 및 패치 내용과 비교 검증"
        ),
        "기댓값": _expected(row, weapon_cond),
    }


def build_tc_cobalt_infusion(row: dict, version: str) -> dict:
    char_raw, item = row["캐릭터"], row["항목"]
    # "2단계 인퓨전 - 라이프 백" → tier="2단계 인퓨전", name="라이프 백"
    if " - " in char_raw:
        tier, inf_name = char_raw.split(" - ", 1)
    else:
        tier, inf_name = "", char_raw
    return {
        "테스트 항목": f"[코발트 인퓨전] {tier} {inf_name} - {item} 확인",
        "사전조건": (
            f"· {version} 패치 적용 완료\n"
            "· 코발트 프로토콜 모드 진행 가능 상태"
        ),
        "테스트 절차": (
            "1. 게임 클라이언트 실행 및 패치 버전 확인\n"
            "2. 코발트 프로토콜 모드 게임 시작\n"
            f"3. 인퓨전 선택 단계({tier})에서 {inf_name} 선택\n"
            f"4. {item} 수치 확인\n"
            "5. 수치가 패치 내용과 일치하는지 검증"
        ),
        "기댓값": _expected(row, inf_name),
    }


BUILDERS = {
    "char_skill":       build_tc_char_skill,
    "char_stat":        build_tc_char_stat,
    "armor":            build_tc_armor,
    "cobalt_mode":      build_tc_cobalt_mode,
    "cobalt_infusion":  build_tc_cobalt_infusion,
}

# TC 생성 대상 유형 — 실험체 스킬/기본 스탯만 포함
INCLUDE_TYPES = frozenset({"char_skill", "char_stat"})

# 유형 정렬 순서 (버전 내 정렬 기준)
TYPE_ORDER = {"char_skill": 0, "char_stat": 1}


# ── Excel 스타일 ───────────────────────────────────────────────────────────────

COL_WIDTHS = {
    "TC ID":     14,
    "테스트 항목":  42,
    "사전조건":    32,
    "테스트 절차":  52,
    "기댓값":     36,
    "결과":       10,
}
COLUMNS = list(COL_WIDTHS.keys())

# 헤더 배경색
HEADER_FILL = PatternFill("solid", fgColor="2E4057")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# 유형별 행 배경색 (연한 색)
TYPE_FILLS = {
    "char_skill":      PatternFill("solid", fgColor="EBF5FB"),  # 연파랑
    "char_stat":       PatternFill("solid", fgColor="EAF4F4"),  # 연청록
    "armor":           PatternFill("solid", fgColor="FEF9E7"),  # 연노랑
    "cobalt_infusion": PatternFill("solid", fgColor="F5EEF8"),  # 연보라
    "cobalt_mode":     PatternFill("solid", fgColor="FDFEFE"),  # 흰색
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")


def style_header(ws):
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22


def write_row(ws, row_idx: int, tc: dict, tc_type: str):
    fill = TYPE_FILLS.get(tc_type, PatternFill())
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=tc.get(col_name, ""))
        cell.fill   = fill
        cell.border = THIN_BORDER
        cell.font   = Font(size=9)
        if col_name == "TC ID":
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.alignment = WRAP

    # 행 높이: 테스트 절차 줄 수 기준
    lines = tc.get("테스트 절차", "").count("\n") + 1
    ws.row_dimensions[row_idx].height = max(15 * lines, 45)


# ── 메인 ──────────────────────────────────────────────────────────────────────

def load_changes(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def generate_tc(rows: list[dict]) -> list[dict]:
    """변경 행 → TC 딕셔너리 리스트 (TC ID 포함)"""
    # 버전별 그룹화 후 유형 순으로 정렬
    from collections import defaultdict
    by_version = defaultdict(list)
    for row in rows:
        by_version[row["패치버전"]].append(row)

    # 버전 정렬: 숫자 기준 내림차순 (최신 패치 먼저)
    def ver_key(v):
        m = re.match(r"(\d+)\.(\d+)([a-z]?)", v)
        if m:
            return (int(m.group(1)), int(m.group(2)), m.group(3) or "")
        return (0, 0, "")

    all_tcs = []
    for version in sorted(by_version, key=ver_key, reverse=True):
        version_rows = by_version[version]
        # 유형 순으로 정렬
        version_rows = [r for r in version_rows if classify(r) in INCLUDE_TYPES]
        version_rows.sort(key=lambda r: (TYPE_ORDER.get(classify(r), 9), r["캐릭터"], r["스킬"]))

        seq = 1
        for row in version_rows:
            tc_type  = classify(row)
            builder  = BUILDERS[tc_type]
            tc_fields = builder(row, version)

            # TC ID: TC-{버전}-{3자리 순번}
            tc_id = f"TC-{version}-{seq:03d}"
            seq  += 1

            all_tcs.append({
                "TC ID":    tc_id,
                "테스트 항목": tc_fields["테스트 항목"],
                "사전조건":   tc_fields["사전조건"],
                "테스트 절차": tc_fields["테스트 절차"],
                "기댓값":    tc_fields["기댓값"],
                "결과":     "",
                "_type":    tc_type,
            })

    return all_tcs


def save_excel(tcs: list[dict], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "패치노트 TC"

    # 헤더
    style_header(ws)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    # 데이터 행
    for row_idx, tc in enumerate(tcs, 2):
        write_row(ws, row_idx, tc, tc["_type"])

    # 행 고정 (헤더)
    ws.freeze_panes = "A2"

    # 범례 시트
    wb_legend = wb.create_sheet("범례")
    legend_data = [
        ("배경색", "유형", "설명"),
        ("연파랑",  "[실험체] 스킬",      "실험체 스킬(Q/W/E/R/P/D) 수치 변경"),
        ("연청록",  "[실험체] 기본 스탯",  "실험체 기본 스탯(체력, 공격력 등) 변경"),
        ("연노랑",  "[방어구]",           "아이템 스탯 변경"),
        ("연보라",  "[코발트 인퓨전]",     "코발트 프로토콜 인퓨전 수치 변경"),
        ("흰색",    "[코발트 모드 보정]",  "코발트 프로토콜 캐릭터별 피해량 보정 변경"),
    ]
    fills = [None, TYPE_FILLS["char_skill"], TYPE_FILLS["char_stat"],
             TYPE_FILLS["armor"], TYPE_FILLS["cobalt_infusion"], TYPE_FILLS["cobalt_mode"]]
    for i, (row_data, fill) in enumerate(zip(legend_data, fills), 1):
        for j, val in enumerate(row_data, 1):
            c = wb_legend.cell(row=i, column=j, value=val)
            if fill:
                c.fill = fill
            if i == 1:
                c.font = Font(bold=True)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
        wb_legend.row_dimensions[i].height = 18
    for col, w in zip("ABC", [10, 20, 40]):
        wb_legend.column_dimensions[col].width = w

    wb.save(path)
    print(f"[저장] {path}  ({len(tcs)}개 TC)")


def main():
    print("=== 패치노트 TC 생성기 ===\n")
    rows = load_changes(INPUT_CSV)
    print(f"변경 데이터 로드: {len(rows)}건")

    tcs = generate_tc(rows)
    print(f"TC 생성: {len(tcs)}건")

    # 유형 분포 출력
    from collections import Counter
    dist = Counter(tc["_type"] for tc in tcs)
    type_labels = {
        "char_skill": "[실험체] 스킬",
        "char_stat":  "[실험체] 기본 스탯",
        "armor":      "[방어구]",
        "cobalt_infusion": "[코발트 인퓨전]",
        "cobalt_mode": "[코발트 모드 보정]",
    }
    for t, cnt in dist.most_common():
        print(f"  {type_labels.get(t, t)}: {cnt}건")
    print()

    save_excel(tcs, OUTPUT_XLS)
    print(f"\n결과: {OUTPUT_XLS.resolve()}")


if __name__ == "__main__":
    main()
